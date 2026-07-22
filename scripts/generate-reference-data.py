from __future__ import annotations

import json
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


ROOT = Path("/workspace/scratch/bb24881553e8/upload")
SITE = Path("/workspace/sites/un-comtrade-builder")
PRODUCTS = ROOT / "2024.01.26 ТНВЭД-6 Десятигрузка (промежуточный) (1).xlsx"
COUNTRIES = Path(os.environ.get(
    "COMTRADE_COUNTRIES_XLSX",
    ROOT / "Справочник стран мира с регионами_ComtradeCode.xlsx",
))
OUTPUT = SITE / "app/data/reference-data.json"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def code(value: object, width: int) -> str:
    text = clean(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(width) if text.isdigit() else text


country_book = openpyxl.load_workbook(COUNTRIES, read_only=True, data_only=True)
country_sheet = country_book.active
country_rows = list(country_sheet.iter_rows(min_row=2, values_only=True))

country_by_comtrade_code: dict[str, dict[str, str]] = {}
region_codes: dict[str, set[str]] = defaultdict(set)
for region, alpha2, alpha3, m49, name_en, name_ru, comtrade_code in country_rows:
    comtrade = clean(comtrade_code)
    if comtrade.endswith(".0"):
        comtrade = comtrade[:-2]
    m49_code = code(m49, 3)
    alpha3_code = clean(alpha3)
    if not comtrade.isdigit() or not alpha3_code:
        continue
    item = {
        "comtradeCode": comtrade,
        "iso": m49_code,
        "alpha2": clean(alpha2),
        "alpha3": alpha3_code,
        "name": clean(name_ru),
        "nameEn": clean(name_en),
    }
    country_by_comtrade_code.setdefault(comtrade, item)
    if clean(region):
        region_codes[clean(region)].add(comtrade)

countries = sorted(
    country_by_comtrade_code.values(),
    key=lambda item: (item["name"].casefold(), int(item["comtradeCode"])),
)
regions = [
    {"name": name, "codes": sorted(codes, key=int)}
    for name, codes in sorted(region_codes.items(), key=lambda pair: pair[0].casefold())
]

product_book = openpyxl.load_workbook(PRODUCTS, read_only=True, data_only=True)
product_sheet = product_book.active
product_rows = list(product_sheet.iter_rows(min_row=2, max_col=6, values_only=True))

hs4_by_code: dict[str, dict[str, str]] = {}
hs6_by_code: dict[str, dict[str, str]] = {}
group_to_hs4: dict[str, set[str]] = defaultdict(set)
subgroup_to_hs4: dict[str, set[str]] = defaultdict(set)
group_to_subgroups: dict[str, set[str]] = defaultdict(set)
hs4_to_hs6: dict[str, set[str]] = defaultdict(set)

for hs4, hs4_desc, hs6, hs6_desc, cargo_group, cargo_subgroup in product_rows:
    hs4_code = code(hs4, 4)
    hs6_code = code(hs6, 6)
    group_name = clean(cargo_group)
    subgroup_name = clean(cargo_subgroup)
    if not re.fullmatch(r"\d{4}", hs4_code) or not re.fullmatch(r"\d{6}", hs6_code):
        continue
    hs4_by_code.setdefault(hs4_code, {"code": hs4_code, "name": clean(hs4_desc)})
    hs6_by_code.setdefault(
        hs6_code,
        {"code": hs6_code, "name": clean(hs6_desc), "parent": hs4_code},
    )
    hs4_to_hs6[hs4_code].add(hs6_code)
    if group_name:
        group_to_hs4[group_name].add(hs4_code)
    if subgroup_name:
        subgroup_to_hs4[subgroup_name].add(hs4_code)
    if group_name and subgroup_name:
        group_to_subgroups[group_name].add(subgroup_name)

payload = {
    "meta": {
        "countriesSource": COUNTRIES.name,
        "productsSource": PRODUCTS.name,
        "countries": len(countries),
        "regions": len(regions),
        "hs4": len(hs4_by_code),
        "hs6": len(hs6_by_code),
    },
    "countries": countries,
    "regions": regions,
    "cargoGroups": sorted(group_to_hs4, key=str.casefold),
    "cargoSubgroups": sorted(subgroup_to_hs4, key=str.casefold),
    "hs4": sorted(hs4_by_code.values(), key=lambda item: item["code"]),
    "hs6": sorted(hs6_by_code.values(), key=lambda item: item["code"]),
    "groupToHs4": {key: sorted(value) for key, value in sorted(group_to_hs4.items())},
    "subgroupToHs4": {key: sorted(value) for key, value in sorted(subgroup_to_hs4.items())},
    "groupToSubgroups": {
        key: sorted(value, key=str.casefold) for key, value in sorted(group_to_subgroups.items())
    },
    "hs4ToHs6": {key: sorted(value) for key, value in sorted(hs4_to_hs6.items())},
}

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
print(json.dumps(payload["meta"], ensure_ascii=False, indent=2))
